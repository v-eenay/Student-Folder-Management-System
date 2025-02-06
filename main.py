import os
import shutil
import openpyxl

# Define paths and files
source_excel_file = 'sample.xlsx'  # Path to the source Excel file to be copied
student_list_file = 'student_list.xlsx'  # Path to the Excel file containing student list and group information
base_folder = 'folders'  # Base folder where all section folders will be created

# Load the Excel workbook and relevant sheets
student_list_wb = openpyxl.load_workbook(student_list_file)
student_list_sheet = student_list_wb['StudentList']  # Sheet containing student details
viva_groups_sheet = student_list_wb['VivaGroups']    # Sheet containing group details

# Logging setup
log_file = 'process_log.txt'
def log_message(message):
    """Logs messages to a text file for tracking purposes."""
    with open(log_file, 'a') as f:
        f.write(message + '\n')
    print(message)

# Clear previous logs if they exist
if os.path.exists(log_file):
    os.remove(log_file)

log_message("=== PROCESS STARTED ===")

# Create a dictionary mapping groups to their respective students (case insensitive)
group_to_students = {}
for row in viva_groups_sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 (skip header)
    _, group, student_name = row
    if group not in group_to_students:
        group_to_students[group] = []
    group_to_students[group].append(student_name.lower())  # Store names in lowercase for comparison

log_message("### Group-to-Students Mapping Created Successfully ###")

# Extract student details from the StudentList sheet
students = []
for row in student_list_sheet.iter_rows(min_row=2, max_col=8, values_only=True):  # Start from row 2
    sno, university_id, name, sem, intake, course, year, section = row  # Adjusted to unpack 8 columns
    if university_id is None or name is None:  # Skip empty rows
        continue
    students.append({
        'university_id': university_id,
        'name': name,
        'section': section
    })

log_message("### Student Details Extracted Successfully ###")

# Ensure base folder exists
os.makedirs(base_folder, exist_ok=True)  # Ensure base folder exists
log_message(f"### Base Folder '{base_folder}' Created/Verified ###")

# Iterate through each student and organize them into respective folders
for student in students:
    university_id = student['university_id']
    name = student['name']
    section = student['section']
    
    # Find the group of the student from the VivaGroups data by comparing names in lowercase
    student_group = None
    for group, group_students in group_to_students.items():
        if name.lower() in group_students:
            student_group = group
            break
    
    if student_group is None:
        log_message(f"### WARNING: No Group Found for Student '{name}' ({university_id}) - Skipping ###")
        continue  # Skip if the student's group cannot be determined
    
    # Create section folder if it doesn't exist
    section_folder = os.path.join(base_folder, section)
    os.makedirs(section_folder, exist_ok=True)
    log_message(f"### Section Folder '{section}' Created/Verified ###")
    
    # Create group folder inside the section folder if it doesn't exist
    group_folder = os.path.join(section_folder, student_group)
    os.makedirs(group_folder, exist_ok=True)
    log_message(f"### Group Folder '{student_group}' Inside Section '{section}' Created/Verified ###")
    
    # Copy the source Excel file into the group folder with the desired filename
    destination_file = os.path.join(group_folder, f"{university_id} {name}.xlsx")
    try:
        shutil.copy(source_excel_file, destination_file)
        log_message(f"### File Copied Successfully for Student '{name}' ({university_id}) ###")
    except Exception as e:
        log_message(f"### ERROR: Failed to Copy File for Student '{name}' ({university_id}): {e} ###")
        continue
    
    # Open the copied Excel file and update the Grading Sheet
    try:
        wb = openpyxl.load_workbook(destination_file)
        grading_sheet = wb['Grading Sheet']
        
        # Write the student name in cell C6 and student ID in cell C7
        grading_sheet['C6'] = name
        grading_sheet['C7'] = university_id
        
        # Save the changes to the Excel file
        wb.save(destination_file)
        log_message(f"### File Updated Successfully for Student '{name}' ({university_id}) ###")
    except Exception as e:
        log_message(f"### ERROR: Failed to Update File for Student '{name}' ({university_id}): {e} ###")

log_message("=== PROCESS COMPLETED ===")